import openpyxl
import xlwings as xw
import json
from tkinter import messagebox
from math import log10, floor

def transform_dictionary(input_dict, transformation_code):
    """
    Transform all keys in a dictionary using the provided transformation code.
    Returns a new dictionary with transformed keys and original values.
    """
    return {
        translate(key, transformation_code): value
        for key, value in input_dict.items()
    }

def try_round_to_sigfigs(num, sig_figs=4, tolerance=1e-2):
    # Rounds a number to specified significant figures, with additional rounding for very close values.
    org_num = num
    try:
        num = float(str(num).strip().replace(',', ''))
    except:
        return org_num

    if num == 0:
        return 0



    # Get the order of magnitude
    magnitude = floor(log10(abs(num)))

    # Calculate the scaling factor
    scale = 10 ** (sig_figs - 1 - magnitude)

    # Round the scaled number
    rounded = round(num * scale) / scale

    # Additional rounding for very close values
    if abs(rounded - round(rounded)) < tolerance:
        rounded = round(rounded)

    return rounded

def find_cells(sheet, search_terms):
    cells = []
    concat_pairs = {}
    concat_cells = {}

    # Parse concat pairs from search terms
    for term in search_terms:
        if '+' in term:
            first, second = term.split('+')
            concat_pairs[first.lower()] = second.lower()

    # Find cells and track concat pairs separately
    for row in sheet.iter_rows():
        for cell in row:
            if not cell.value:
                continue
            cell_value = str(cell.value).lower()

            if cell_value in [term.lower() for term in search_terms if '+' not in term]:
                cells.append(cell)

            if cell_value in concat_pairs:
                second_value = concat_pairs[cell_value]
                for second_cell in row:
                    if second_cell.value and str(second_cell.value).lower() == second_value:
                        concat_cells[cell] = second_cell
                        cells.append(cell)
                        break

    return cells, concat_cells


def process_sheet(sheet, headers):
    header_cells, concat_pairs = find_cells(sheet, headers)
    tables = []

    for cell in header_cells:
        table_length = get_table_length_rows(sheet, cell)
        left_col, right_col = get_table_cols(sheet, cell)

        table_dict_list = table_to_list(sheet, cell.row, left_col, cell.row + table_length, right_col)

        original_header = next(h for h in headers if h.lower() == cell.value.lower() or
                               ('+' in h and h.split('+')[0].lower() == cell.value.lower()))

        tag_dict = list_to_tag_dict(table_dict_list, original_header)
        tables.append(tag_dict)

    return tables


def list_to_tag_dict(table_list, tag):
    tag_dict = {}
    duplicate_counts = {}  # Track how many times we've seen each tag key

    for row in table_list:
        if '+' in tag:
            first_tag, second_tag = tag.split('+')
            if first_tag in row and second_tag in row:
                base_tag_key = str(row[first_tag]) + str(row[second_tag])
        else:
            if tag in row:
                base_tag_key = row[tag]

        if base_tag_key in tag_dict:
            # If we've seen this tag before, get the next available number
            if base_tag_key not in duplicate_counts:
                duplicate_counts[base_tag_key] = 1

            duplicate_counts[base_tag_key] += 1
            tag_key = f"{base_tag_key}_{duplicate_counts[base_tag_key]}"
        else:
            tag_key = base_tag_key

        tag_dict[tag_key] = row

    return tag_dict


def find_cells_old(sheet, search_terms):
    cells = []
    # Find the row in the sheet that contains any of the specified search terms
    for term in search_terms:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and term.lower() == str(cell.value).lower():
                    cells.append(cell)
    return cells

def combine_tables(tables):
    if tables:
        tag_dict_combined = tables[0].copy()
    else:
        print('Empty tables')
        return

    for table in tables[1:]:
        for tag in table:
            if table[tag] is not None:
                # Use setdefault to ensure the tag exists in tag_dict_combined
                try:
                    tag_dict_combined[tag].update(table[tag])
                except Exception as e:
                    print('ERROR: ',e)
            else:
                print(f'Error: No match found for {tag}')
    print(tag_dict_combined)
    return tag_dict_combined


def get_unique_sheet_name(datasheet, ds_prefix, sheet_number):
    """Generate unique sheet name, incrementing suffix if needed"""
    base_name = f"{ds_prefix}{str(sheet_number).zfill(2)}"
    name = base_name
    suffix = 1

    while name in datasheet.sheets:
        name = f"{base_name}_{suffix}"
        suffix += 1

    return name

def add_datasheets(datasheet, source_sheet_name, tag_cell_values, datasheet_coord, ds_prefix,
                   rows_per_sheet=1, custom_sort=None, key_coordinate='I12'):
    """
    Manages Excel sheets by adding or updating data based on tags.

    Args:
        datasheet: Excel workbook object
        source_sheet_name: Name of template sheet to copy from
        tag_cell_values: Dict of tags mapping to cell value updates {tag: {cell_ref: value}}
        datasheet_coord: Cell reference where sheet name should be written
        ds_prefix: Prefix for new sheet names
        rows_per_sheet: Number of data rows per sheet (default=1)
        custom_sort: Optional function for custom tag sorting
        key_coordinate: Starting cell reference for tag placement (default='I12')
    """
    source_sheet = datasheet.sheets[source_sheet_name]
    added_sheets = set()  # Track sheets we modify

    # Build dictionary of existing tags and their locations in workbook
    existing_tags = {}
    for sheet in datasheet.sheets:
        if sheet.name.startswith(ds_prefix) or ds_prefix == '':
            for i in range(rows_per_sheet):
                offset_coord = increment_cell_reference(key_coordinate, i)
                tag_value = sheet.range(offset_coord).value
                if tag_value:
                    existing_tags[tag_value] = (sheet.name, offset_coord)


    print(f'Existing tags: {existing_tags}')

    # Sort tags according to custom function or default alphabetical
    sorted_keys = sorted(tag_cell_values, key=custom_sort) if custom_sort else sorted(tag_cell_values)
    print(f'Tags to process: {sorted_keys}')

    # Process each tag - either update existing or create new entry
    count = len(existing_tags)
    for tag in sorted_keys:
        if tag in existing_tags:
            # Update values for existing tag
            print(f'Updating existing tag {tag}')
            sheet_name, tag_coord = existing_tags[tag]
            target_sheet = datasheet.sheets[sheet_name]
        else:
            # Create new sheet if needed based on rows_per_sheet
            print(f'Adding new tag {tag}')
            if count % rows_per_sheet == 0:
                sheet_name = get_unique_sheet_name(datasheet, ds_prefix, (count // rows_per_sheet) + 1)
                try:
                    datasheet.sheets[sheet_name].delete()  # Remove if exists
                except:
                    pass
                target_sheet = source_sheet.copy(name=sheet_name)
                added_sheets.add(sheet_name)  # Track new sheets
                target_sheet.range(datasheet_coord).value = sheet_name
            else:
                sheet_name = get_unique_sheet_name(datasheet, ds_prefix, (count // rows_per_sheet) + 1)
                print('Might need to make sure existing sheet names dont cause issues')
                target_sheet = datasheet.sheets[sheet_name]

            tag_coord = increment_cell_reference(key_coordinate, count % rows_per_sheet)
            count += 1

        # Update all specified cells for this tag
        cell_values = tag_cell_values[tag]
        row_offset = int(tag_coord[1:]) - int(key_coordinate[1:])

        for cell, value in cell_values.items():
            try:
                target_cell = increment_cell_reference(cell, row_offset)
                value = try_round_to_sigfigs(value)
                target_sheet.range(target_cell).value = value
            except Exception as e:
                print(f"Error updating cell {cell} for tag {tag}: {e}")

    return list(added_sheets)  # Return list of all sheets we Added

def translate(input_string, transformation_code):
    try:
        x_fn = eval(f'lambda x: {transformation_code}')
        transformed_string = x_fn(input_string)
        return transformed_string
    except Exception as e:
        print(f'error applying transformation: {e}')
        return input_string


def update_datasheets(datasheet_path, tag_cell_values, ds_prefix, rows_per_sheet=1, key_coordinate='I12'):

    datasheet = xw.Book(datasheet_path)

    # Get existing tags and their locations
    # here were baiscally machine-gunning the datasheet with data
    for sheet in datasheet.sheets:
        if sheet.name.startswith(ds_prefix) or ds_prefix == '':
            for i in range(rows_per_sheet):
                tag_coord = increment_cell_reference(key_coordinate, i)
                tag_value = sheet.range(tag_coord).value
                if tag_value in tag_cell_values:# and re.search(tag_pattern, str(tag_value)):
                    coord_values = tag_cell_values[tag_value]
                    # so coord values is going to be like:
                    #{'C38': 'LE-3101', 'E38': 'AT1-2000-PR-PID-111', 'F38': 'TK-3101A', 'H38': 'SOFTENING FEED TANK A'}
                    for coord, value in coord_values.items():
                        current_coord = increment_cell_reference(coord,i)
                        sheet.range(current_coord).value = value

                    #we need to update the values here and not later

def increment_cell_reference(cell_ref, increment):
    col = ''.join(filter(str.isalpha, cell_ref))
    row = int(''.join(filter(str.isdigit, cell_ref)))
    return f"{col}{row + increment}"

def generate_dictionary_from_xlsx(wb_path, headers):
    wb = openpyxl.load_workbook(wb_path, read_only=False, data_only=True)
    # process_conditions['09-PM-006-2'] = ...
    wb_tag_data = {}
    sheet_names = wb.sheetnames

    for i, sheet in enumerate(wb.worksheets):
        print(f'processing {sheet_names[i]}')

        tag_tables = process_sheet(sheet, headers)

        sheet_data = combine_tables(tag_tables)

        if sheet_data:
            wb_tag_data.update(sheet_data)

    return wb_tag_data

def get_value_above(sheet, row, col):
    """Get value from merged cell above or regular cell."""
    current_row = row
    while current_row > 1:
        current_row -= 1
        for range_string in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_string.bounds
            if (min_row <= current_row <= max_row and
                min_col <= col <= max_col):
                print('mereged cell ', sheet.cell(min_row, min_col).value)
                return sheet.cell(min_row, min_col).value
        cell_value = sheet.cell(row=current_row, column=col).value
        if cell_value is not None:
            return cell_value
    return None

def table_to_list(sheet, start_row, start_col, end_row, end_col, column_headers=False):
    matrix = []
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row,
                             min_col=start_col, max_col=end_col,
                             values_only=False):
        row_values = [cell.value for cell in row]
        matrix.append(row_values)

    headers = matrix.pop(0)
    modified_headers = headers.copy()

    # Handle duplicates considering merged cells
    header_counts = {}
    for i, header in enumerate(headers):
        if header is None:
            continue
        header_counts[header] = header_counts.get(header, 0) + 1

    for i, header in enumerate(headers):
        if header_counts[header] > 1:
            col_index = start_col + i
            print(sheet.cell(start_row-1, col_index).value)
            above_value = get_value_above(sheet, start_row, col_index)
            if above_value:
                modified_headers[i] = f"{above_value}_{header}"

    table_list = []
    for row in matrix:
        row_dict = dict(zip(modified_headers, row))
        table_list.append(row_dict)

    return table_list

def get_table_length_rows(sheet, header_cell):
    start_row = header_cell.row + 1
    end_value = header_cell.value
    max_row = sheet.max_row
    current_row = start_row
    # Initialize table_length to 0
    table_length = 0

    # Use iter_rows to iterate over rows starting from the row after the header
    for row in sheet.iter_rows(min_row=start_row, values_only=True):
        current_value = row[header_cell.column - 1]  # Adjusting for zero-based indexing

        # Break the loop if an empty cell is encountered or if the current value matches the end value
        if current_row >= max_row or current_value == end_value:
            break

        current_row += 1
        table_length += 1

    return table_length

def get_table_cols(sheet, header_cell):
    start_col = header_cell.column
    end_value = header_cell.value
    row = header_cell.row

    # Find the leftmost column
    left_col = start_col
    for col in range(start_col - 1, 0, -1):
        if sheet.cell(row=row, column=col).value is None:
            break
        left_col = col

    # Find the rightmost column
    right_col = start_col
    for col in range(start_col + 1, sheet.max_column + 1):
        cell_value = sheet.cell(row=row, column=col).value
        if cell_value is None or cell_value == end_value:
            right_col = col - 1
            break
        right_col = col

    return left_col, right_col

def load_dict_from_json(file_path):
    """
    Load a dictionary from a JSON file.

    Args:
    file_path (str): The path to the JSON file.

    Returns:
    dict: The dictionary loaded from the JSON file.

    Raises:
    FileNotFoundError: If the specified file is not found.
    json.JSONDecodeError: If the file is not valid JSON.
    """
    try:
        with open(file_path, 'r') as json_file:
            return json.load(json_file)
    except FileNotFoundError:
        raise FileNotFoundError(f"The file {file_path} was not found.")
    except json.JSONDecodeError:
        raise json.JSONDecodeError(f"The file {file_path} is not valid JSON.")


def analyze_nested_dict_keys(dict_of_dicts):
    if not dict_of_dicts:
        return {
            'consistent': True,
            'inconsistencies': []
        }

    # Get first dictionary's keys as reference
    reference_keys = set(next(iter(dict_of_dicts.values())).keys())

    inconsistencies = []

    # Check each dictionary against the reference
    for key, d in dict_of_dicts.items():
        current_keys = set(d.keys())
        if current_keys != reference_keys:
            missing_keys = reference_keys - current_keys
            extra_keys = current_keys - reference_keys

            inconsistency = {
                'key': key,  # Using the dictionary key instead of index
                'missing_keys': list(missing_keys) if missing_keys else None,
                'extra_keys': list(extra_keys) if extra_keys else None
            }
            inconsistencies.append(inconsistency)

    return {
        'consistent': len(inconsistencies) == 0,
        'reference_keys': list(reference_keys),
        'total_dictionaries': len(dict_of_dicts),
        'inconsistent_count': len(inconsistencies),
        'inconsistencies': inconsistencies
    }


def show_nested_dict_analysis(dict_of_dicts):
    result = analyze_nested_dict_keys(dict_of_dicts)

    # Build the message string
    if result['consistent']:
        message = "All nested dictionaries have the same keys!"
    else:
        message = f"Found {result['inconsistent_count']} inconsistent dictionaries out of {result['total_dictionaries']}\n\n"
        message += f"Reference keys: {', '.join(result['reference_keys'])}\n\n"
        message += "Inconsistencies:\n"

        for item in result['inconsistencies']:
            message += f"\nDictionary with key '{item['key']}':\n"
            if item['missing_keys']:
                message += f"  Missing keys: {', '.join(item['missing_keys'])}\n"
            if item['extra_keys']:
                message += f"  Extra keys: {', '.join(item['extra_keys'])}\n"

    messagebox.showinfo("Nested Dictionary Key Analysis Results", message)

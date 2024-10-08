import tkinter as tk
from tkinter import messagebox, filedialog
import re
import openpyxl
import ast
import json
import os
from coords_to_fields import CoordsToFieldsGenerator


def extract_data_from_datasheets(file_path, init_tag_coord, init_coords_to_fields, tags_per_sheet=1):
    """
     Extracts data from Excel datasheets using openpyxl.

     This function loads an Excel workbook from the given file path and iterates through its worksheets,
     extracting data based on initial coordinates and mappings between coordinates and field names.
     It returns a dictionary where keys are tags (e.g., sheet names) and values are dictionaries containing
     field names as keys and their corresponding cell values as values.
     """
    # Load the workbook
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    # Initialize an empty dictionary to store the extracted data
    all_tag_data = {}

    for ws in wb.worksheets:
        tag_coord = init_tag_coord
        coords_to_fields = init_coords_to_fields
        # Iterate over each coordinate-field name pair
        for i in range(tags_per_sheet):
            tag = ws[tag_coord].value
            if not tag:
                continue
            tag_data = {}
            for coord, field_name in coords_to_fields.items():
                # Extract the value from the specified cell
                cell_value = ws[coord].value
                # Add the field name and its corresponding value to the extracted_data dictionary
                tag_data[field_name] = cell_value

            # add tag to all tag data
            all_tag_data[tag] = tag_data
            # increment the coords and tag
            coords_to_fields = increment_coords_to_fields(coords_to_fields)
            tag_coord = increment_coord(tag_coord)

    # Return the dictionary of extracted data
    return all_tag_data


def increment_coords_to_fields(coords_to_fields):
    new_coords_to_fields = {}
    for coord, field_name in coords_to_fields.items():
        new_coord = increment_coord(coord)
        new_coords_to_fields[new_coord] = field_name

    return new_coords_to_fields


def increment_coord(coord):
    letter_part, numeric_part = split_text_on_first_number(coord)
    new_coord = letter_part + str(int(numeric_part) + 1)
    return new_coord


def split_text_on_first_number(text):
    match = re.search(r'\d', text)  # Search for the first digit in the text
    if match:
        first_number_index = match.start()  # Get the index of the first digit
        text_before_number = text[:first_number_index]  # Text before the first number
        text_after_number = text[first_number_index:]  # Text from the first number onward
        return text_before_number, text_after_number
    else:
        # If there is no digit in the text, return the entire text
        return text, ''


class DatasheetExtractor:
    def __init__(self, root, callback=None):
        self.root = root
        self.root.title("Datasheet Data Extraction GUI")
        self.callback = callback

        self.create_widgets()
        self.create_top_menu()

    def create_widgets(self):
        # File path entry
        tk.Label(self.root, text="File Path").grid(row=0)
        self.file_path_entry = tk.Entry(self.root)
        self.file_path_entry.grid(row=0, column=1)

        browse_button = tk.Button(self.root, text="Browse", command=self.browse_file_path)
        browse_button.grid(row=0, column=2)

        # Init Tag Coord entry
        tk.Label(self.root, text="Init Tag Coord").grid(row=1)
        self.init_tag_coord_entry = tk.Entry(self.root)
        self.init_tag_coord_entry.grid(row=1, column=1)



        # Init Coords to Fields entry
        tk.Label(self.root, text="Init Coords to Fields").grid(row=2)
        self.init_coords_to_fields_entry = tk.Entry(self.root)
        self.init_coords_to_fields_entry.grid(row=2, column=1)

        set_coords_button = tk.Button(self.root, text="Set Coordinate Fields", command=self.generate_coords_to_fields)
        set_coords_button.grid(row=2, column=2)

        # Tags per Sheet entry
        tk.Label(self.root, text="Tags per Sheet").grid(row=3)
        self.tags_per_sheet_entry = tk.Entry(self.root)
        self.tags_per_sheet_entry.grid(row=3, column=1)

        # Start extraction button
        start_button = tk.Button(self.root, text="Start Extraction", command=self.start_extraction)
        start_button.grid(row=4, columnspan=3)

        # Clear entries button
        clear_button = tk.Button(self.root, text="Clear Entries", command=self.clear_entries)
        clear_button.grid(row=5, columnspan=3)

    def create_top_menu(self):
        menu_bar = tk.Menu(self.root)
        self.root.config(menu=menu_bar)

        # File menu
        file_menu = tk.Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="Open Datasheet", command=self.browse_file_path)
        file_menu.add_command(label="Save", command=self.dummy_command)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.root.quit)

        # Edit menu
        edit_menu = tk.Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="Edit", menu=edit_menu)
        edit_menu.add_command(label="Edit Datasheet", command=self.dummy_command)
        edit_menu.add_command(label="Copy", command=self.dummy_command)
        edit_menu.add_command(label="Paste", command=self.dummy_command)

        # View menu
        view_menu = tk.Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="View", menu=view_menu)
        view_menu.add_command(label="View Datasheet", command=self.dummy_command)
        view_menu.add_command(label="Zoom Out", command=self.dummy_command)

        # Commands menu
        commands_menu = tk.Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="Commands", menu=commands_menu)
        commands_menu.add_command(label="Set Coordinates Fields", command=self.generate_coords_to_fields)

        # Help menu
        help_menu = tk.Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="Help", menu=help_menu)
        help_menu.add_command(label="About", command=self.show_help)

    def browse_file_path(self):
        filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if filepath:
            self.file_path_entry.delete(0, tk.END)
            self.file_path_entry.insert(0, filepath)

    def start_extraction(self):
        file_path = self.file_path_entry.get()
        init_tag_coord = self.init_tag_coord_entry.get()
        init_coords_to_fields = dict(eval(self.init_coords_to_fields_entry.get()))
        tags_per_sheet = int(self.tags_per_sheet_entry.get() if self.tags_per_sheet_entry.get().isdigit() else 1)

        try:
            result = extract_data_from_datasheets(file_path, init_tag_coord, init_coords_to_fields, tags_per_sheet)
            messagebox.showinfo("Result", "Data extraction completed successfully.")

            save_path = filedialog.asksaveasfilename(defaultextension=".json",
                                                     filetypes=[("JSON files", "*.json")])

            if save_path:
                with open(save_path, 'w') as json_file:
                    json.dump(result, json_file, indent=4)
                messagebox.showinfo("Save Successful", f"Result saved to {save_path}")

                if messagebox.askyesno("Open File", "Do you want to open the saved file?"):
                    os.startfile(save_path)


            else:
                messagebox.showinfo("Save Cancelled", "Result was not saved.")

            if self.callback:
                self.callback(result)

        except Exception as e:
            messagebox.showerror("Error", str(e))

    def clear_entries(self):
        self.file_path_entry.delete(0, tk.END)
        self.init_tag_coord_entry.delete(0, tk.END)
        self.init_coords_to_fields_entry.delete(0, tk.END)
        self.tags_per_sheet_entry.delete(0, tk.END)

    def generate_coords_to_fields(self):
        xlsx_path = self.file_path_entry.get()
        if not xlsx_path:
            messagebox.showerror("Error", "Please select an Excel file first.")
            return

        try:
            init_coords_to_fields_str = self.init_coords_to_fields_entry.get()
            initial_coords_dict = ast.literal_eval(init_coords_to_fields_str) if init_coords_to_fields_str else {}
        except:
            messagebox.showwarning("Warning",
                                   "Failed to parse Init Coords to Fields. Starting with an empty dictionary.")
            initial_coords_dict = {}

        generator = CoordsToFieldsGenerator(self.root, xlsx_path, initial_coords_dict)
        generator.generate()

        # Wait for the coords_window to be destroyed
        self.root.wait_window(generator.coords_window)

        # Get the result and update the entry
        result = generator.get_result()
        self.init_coords_to_fields_entry.delete(0, tk.END)
        self.init_coords_to_fields_entry.insert(0, str(result))

    def dummy_command(self):
        print("Menu item clicked!")

    def show_help(self):
        messagebox.showinfo("About Data Extraction GUI",
                            """Data Extraction GUI
                            Version: 1.0
                            Author: Your Name
                            This application allows users to extract data from Excel datasheets.""")


'''
def on_extraction_complete(result, save_path):
    print(f"Extraction complete. Result saved to: {save_path}")

root = tk.Tk()
app = DatasheetExtractor(root, callback=on_extraction_complete)
root.mainloop()
'''
